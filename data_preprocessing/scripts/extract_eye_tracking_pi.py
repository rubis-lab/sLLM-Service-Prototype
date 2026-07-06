
"""
260123 수정: 엑셀 시트에서 pi 목록들 불러와서 pi 작성하도록 수정. 아직 AOI, similarity 관련 계산 방법 반영 x -> null로 찍힘.

extract_eye_pi.py

Compute eye-tracking Primitive Indicators (PIs) from a CSV of gaze samples and
export them to a JSON windowed format.

- Input CSV (required columns):
    timestamp : int/float ms since epoch (or ISO8601 string)
    x         : gaze x in pixels
    y         : gaze y in pixels

- Windowing:
    Sliding windows with user-defined window_size (default 60s) and stride (default 30s).

- PI naming:
    Each PI key is the Excel "Name" lowercased + snake_cased, with "_{mean|std|min|max}" suffix.

Notes / assumptions:
- Fixation vs saccade segmentation uses a simple I-VT velocity threshold on on-screen points.
  Threshold is adaptive: max(BASE_SACCADE_THR_PX_S, median(speed) + MAD_MULT * MAD(speed)).
- Off-screen = (x,y) outside the screen bounds or NaN.
  Most spatial/kinematic features are computed on on-screen samples only;
  on/off-screen ratios are computed using dwell-time.
- Edge / center region definitions are parameterized (EDGE_MARGIN_PCT, CENTER_REGION_PCT).
- "Backtrack saccade" is interpreted as direction reversal: |Δθ - π| < BACKTRACK_TAU_RAD.
- Saccade Direction stats use circular mean/std (angles are periodic).
- IMPORTANT (your request): if a PI yields only ONE value in a window,
  mean/min/max are filled with that value and std is set to None.

Dependencies:
  numpy, pandas
  (optional) scipy for convex hull. If scipy is missing, hull area/perimeter become 0.

Example:
    python data_preprocessing/scripts/extract_eye_tracking_pi.py \
    --base_path pdss_data \
    --output_base_path data_preprocessing/primitive_indicator \
    --window_size_sec 2 \
    --stride_sec 2 \
    --screen_width_px 1440 \
    --screen_height_px 900
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from xml.etree import ElementTree as ET

import numpy as np
import pandas as pd
from pandas.errors import EmptyDataError

try:
    from scipy.spatial import ConvexHull  # type: ignore
except Exception:
    ConvexHull = None  # type: ignore

try:
    # Python 3.9+
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None  # type: ignore

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_PREPROCESSING_DIR = PROJECT_ROOT / "data_preprocessing"
DEFAULT_RAW_DATA_DIR = PROJECT_ROOT / "pdss_data"
DEFAULT_PI_OUTPUT_DIR = DATA_PREPROCESSING_DIR / "primitive_indicator"

# =========================
# Task time ranges (start_t / end_t) support
# =========================
def _ts_to_ms_utc(ts_str: str) -> Optional[int]:
    if ts_str is None:
        return None
    t = pd.to_datetime(ts_str, errors="coerce", utc=True)
    if pd.isna(t):
        return None
    return int(t.value // 1_000_000)  # ns -> ms


def load_task_time_index(task_time_json: str) -> Dict[Tuple[str, str], Tuple[int, int]]:
    """
    Load task start/end timestamps from JSON created by the result-based script.

    Expected (preferred) schema:
      {
        "items": [
          {"student_id" or "client_id": "...", "task": "...", "status":"ok",
           "start_t":"ISO", "end_t":"ISO", ...},
          ...
        ]
      }

    Returns:
      dict[(client_id, task)] = (start_ms, end_ms)
    """
    with open(task_time_json, "r", encoding="utf-8") as f:
        obj = json.load(f)

    items = obj.get("items", [])
    idx: Dict[Tuple[str, str], Tuple[int, int]] = {}

    for it in items:
        if not isinstance(it, dict):
            continue
        if it.get("status") != "ok":
            continue

        cid = it.get("client_id") or it.get("student_id")
        task = it.get("task")
        if not cid or not task:
            continue

        s_ms = _ts_to_ms_utc(it.get("start_t"))
        e_ms = _ts_to_ms_utc(it.get("end_t"))
        if s_ms is None or e_ms is None:
            continue
        if e_ms <= s_ms:
            continue

        idx[(str(cid), str(task))] = (int(s_ms), int(e_ms))

    return idx


# =========================
# Parameters (tune as needed)
# =========================
DEFAULT_SCREEN_WIDTH = 1440
DEFAULT_SCREEN_HEIGHT = 900

# Windowing
DEFAULT_WINDOW_SIZE_S = 2.0
DEFAULT_STRIDE_S = 2.0

# Gaps / dwell time handling
MAX_GAP_MS = 200.0       # for velocity / segmentation (ignore intervals larger than this)
MAX_DWELL_MS = 200.0     # for region duration accumulation (ignore intervals larger than this)

# Fixation / saccade segmentation (I-VT)
MIN_FIX_MS = 100.0
MIN_SAC_MS = 15.0
BASE_SACCADE_THR_PX_S = 3000.0
MAD_MULT = 6.0

# Entropy / RQA discretization
ENTROPY_BIN_PX = 50.0
RQA_BIN_PX = 50.0
RQA_LMIN = 2
MAX_RQA_POINTS = 2000

# Regions (edge/center)
EDGE_MARGIN_PCT = 0.10
CENTER_REGION_PCT = 0.50  # center rectangle width/height is this fraction of screen

# AOI defaults for gng/flanker/dnb
DEFAULT_AOI_X_MIN = 510.0
DEFAULT_AOI_X_MAX = 930.0
DEFAULT_AOI_Y_MIN = 240.0
DEFAULT_AOI_Y_MAX = 660.0
DEFAULT_AOI_TASKS = ("gng", "flanker", "dnb")

# Flanker distractor layout from temp_260513/0610/add_flanker_aoi_pi_2_2.py.
FLANKER_SCREEN_WIDTH = 1440.0
FLANKER_SCREEN_HEIGHT = 900.0
FLANKER_ARROW_SIZE = 248.0
FLANKER_ARROW_GAP = 48.0
FLANKER_ARROW_COUNT = 5

# Direction-related
SACCADE_AXIS_THR_RAD = math.radians(15.0)  # horizontal/vertical classification
BACKTRACK_TAU_RAD = math.radians(20.0)     # for "backtrack" definition


# =========================
# Task-specific window size & stride
# =========================
# task 폴더명(<task>) 기준으로 window_size_sec과 stride_sec를 다르게 설정
# TASK_WINDOW_SIZE_SEC: Dict[str, float] = {
#     "vst": 10.0, #다 80으로 dnb 까지
#     "ast": 10.0,
#     "gng": 10.0,
#     "flanker": 10.0,
#     "dnb": 10.0,
#     "rocf_copy_low": 14.0,
#     "rocf_recall_low": 21.0,
#     "rocf_copy_mid": 16.0,
#     "rocf_recall_mid": 24.0,
#     "rocf_copy_high": 20.0,
#     "rocf_recall_high": 30.0,
# }

# TASK_STRIDE_SEC: Dict[str, float] = {
#     "vst": 2.0, #다 2로 dnb 까지
#     "ast": 2.0,
#     "gng": 2.0,
#     "flanker": 2.0,
#     "dnb": 2.0,
#     "rocf_copy_low": 14.0,
#     "rocf_recall_low": 21.0,
#     "rocf_copy_mid": 16.0,
#     "rocf_recall_mid": 24.0,
#     "rocf_copy_high": 20.0,
#     "rocf_recall_high": 30.0,
# }

TARGET_TASK = ["vst", "ast", "gng", "flanker", "dnb"]

# TASK_WINDOW_SIZE_SEC: Dict[str, float] = {
#     "vst": 60.0,
#     "ast": 40.0,
#     "gng": 40.0,
#     "flanker": 40.0,
#     "dnb": 75.0,
#     "rocf_copy_low": 14.0,
#     "rocf_recall_low": 21.0,
#     "rocf_copy_mid": 16.0,
#     "rocf_recall_mid": 24.0,
#     "rocf_copy_high": 20.0,
#     "rocf_recall_high": 30.0,
# }

# TASK_STRIDE_SEC: Dict[str, float] = {
#     "vst": 60.0,
#     "ast": 40.0,
#     "gng": 40.0,
#     "flanker": 40.0,
#     "dnb": 75.0,
#     "rocf_copy_low": 14.0,
#     "rocf_recall_low": 21.0,
#     "rocf_copy_mid": 16.0,
#     "rocf_recall_mid": 24.0,
#     "rocf_copy_high": 20.0,
#     "rocf_recall_high": 30.0,
# }

def resolve_window_size_for_task(task: str, default_window_size: float) -> float:
    mapping = globals().get("TASK_WINDOW_SIZE_SEC", None)
    if task is None or not isinstance(mapping, dict):
        return float(default_window_size)
    return float(TASK_WINDOW_SIZE_SEC.get(str(task).lower(), default_window_size))
    
def resolve_stride_for_task(task: str, default_stride: float) -> float:
    mapping = globals().get("TASK_STRIDE_SEC", None)
    if task is None or not isinstance(mapping, dict):
        return float(default_stride)
    return float(TASK_STRIDE_SEC.get(str(task).lower(), default_stride))

def fmt_sec(x: float) -> str:
    # 80.0 -> "80", 14.0 -> "14", 12.5 -> "12p5" (파일명 안전)
    if float(x).is_integer():
        return str(int(x))
    return str(x).replace(".", "p")



STAT_SUFFIXES = ("mean", "std", "min", "max")

def _xlsx_col_to_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", str(cell_ref).upper())
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1

def _xlsx_cell_value(cell: ET.Element, shared_strings: List[str], ns: Dict[str, str]) -> Optional[str]:
    value_el = cell.find("x:v", ns)
    inline_el = cell.find("x:is/x:t", ns)
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr" and inline_el is not None:
        return inline_el.text
    if value_el is None:
        return None
    raw = value_el.text
    if raw is None:
        return None
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except Exception:
            return None
    return raw

def _load_name_column_from_xlsx(excel_path: str, sheet_name: str) -> List[str]:
    ns = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(excel_path) as zf:
        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("x:si", ns):
                texts = [t.text or "" for t in si.findall(".//x:t", ns)]
                shared_strings.append("".join(texts))

        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("rel:Relationship", ns)
            if "Id" in rel.attrib and "Target" in rel.attrib
        }

        sheet_rel_id = None
        for sheet in workbook.findall("x:sheets/x:sheet", ns):
            if sheet.attrib.get("name") == sheet_name:
                sheet_rel_id = sheet.attrib.get(f"{{{ns['r']}}}id")
                break
        if sheet_rel_id is None or sheet_rel_id not in rel_targets:
            raise ValueError(f"[{excel_path}] Missing sheet: {sheet_name}")

        target = rel_targets[sheet_rel_id]
        sheet_path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
        sheet_root = ET.fromstring(zf.read(sheet_path))
        rows = sheet_root.findall("x:sheetData/x:row", ns)
        if not rows:
            raise ValueError(f"[{excel_path}/{sheet_name}] Empty sheet")

        header_cells: Dict[int, str] = {}
        for cell in rows[0].findall("x:c", ns):
            ref = cell.attrib.get("r", "")
            value = _xlsx_cell_value(cell, shared_strings, ns)
            if value is not None:
                header_cells[_xlsx_col_to_index(ref)] = value
        if "Name" not in header_cells.values():
            raise ValueError(f"[{excel_path}/{sheet_name}] Missing column: Name")
        name_idx = next(idx for idx, value in header_cells.items() if value == "Name")

        names: List[str] = []
        for row in rows[1:]:
            values: Dict[int, str] = {}
            for cell in row.findall("x:c", ns):
                ref = cell.attrib.get("r", "")
                value = _xlsx_cell_value(cell, shared_strings, ns)
                if value is not None:
                    values[_xlsx_col_to_index(ref)] = value
            if name_idx in values and str(values[name_idx]).strip():
                names.append(str(values[name_idx]).strip())
        return names

def load_pi_fullnames_from_excel(excel_path: str, sheet_name: str = "eye-tracking") -> List[str]:
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        if "Name" not in df.columns:
            raise ValueError(f"[{excel_path}/{sheet_name}] Missing column: Name")
        names = df["Name"].dropna().astype(str).tolist()
    except Exception:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"[{excel_path}] Missing sheet: {sheet_name}")
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None or "Name" not in header:
                raise ValueError(f"[{excel_path}/{sheet_name}] Missing column: Name")
            name_idx = list(header).index("Name")
            names = [str(row[name_idx]) for row in rows if row and row[name_idx] is not None]
        except Exception:
            names = _load_name_column_from_xlsx(excel_path, sheet_name)

    # 중복 제거(순서 유지)
    seen = set()
    out = []
    for n in names:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def split_fullname(fullname: str) -> Tuple[str, str]:
    """
    fullname: et_xxx_mean 형태
    return: (base_full, stat)  -> (et_xxx, mean)
    """
    for s in STAT_SUFFIXES:
        suf = "_" + s
        if fullname.endswith(suf):
            return fullname[: -len(suf)], s
    # 형식이 다르면 일단 base/stat을 못 나누니 (stat=None 처리 대신) 그냥 에러로 두는 게 안전
    raise ValueError(f"PI name does not end with a valid stat suffix: {fullname}")



# =========================
# Helpers
# =========================
def keyify(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def angular_diff(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    """Smallest absolute angular difference (radians) between angles a and b."""
    d = (a - b + np.pi) % (2 * np.pi) - np.pi
    return np.abs(d)


def circular_mean(angles: np.ndarray) -> float:
    s = float(np.mean(np.sin(angles)))
    c = float(np.mean(np.cos(angles)))
    return float(math.atan2(s, c))


def circular_std(angles: np.ndarray) -> Optional[float]:
    angles = np.asarray(angles, dtype=float)
    angles = angles[np.isfinite(angles)]
    if angles.size <= 1:
        return None
    s = float(np.mean(np.sin(angles)))
    c = float(np.mean(np.cos(angles)))
    R = math.sqrt(s * s + c * c)
    if R <= 0:
        return float(np.pi / math.sqrt(3))
    return float(math.sqrt(-2.0 * math.log(R)))


def stats_from_vec(vec: np.ndarray, circular: bool = False) -> Dict[str, Optional[float]]:
    """
    If vec has only ONE value:
      mean=min=max=value, std=None  (per your request)
    """
    vec = np.asarray(vec, dtype=float)
    vec = vec[np.isfinite(vec)]
    if vec.size == 0:
        return {"mean": None, "std": None, "min": None, "max": None}

    if circular:
        # normalize to [-pi, pi]
        v = ((vec + np.pi) % (2 * np.pi)) - np.pi
        if v.size == 1:
            val = float(v[0])
            return {"mean": val, "std": None, "min": val, "max": val}
        m = circular_mean(v)
        sd = circular_std(v)
        return {"mean": float(m), "std": (None if sd is None else float(sd)),
                "min": float(np.min(v)), "max": float(np.max(v))}

    if vec.size == 1:
        val = float(vec[0])
        return {"mean": val, "std": None, "min": val, "max": val}

    m = float(np.mean(vec))
    sd = float(np.std(vec, ddof=1))
    return {"mean": m, "std": sd, "min": float(np.min(vec)), "max": float(np.max(vec))}


def entropy_2d(x: np.ndarray, y: np.ndarray, bin_px: float = ENTROPY_BIN_PX) -> float:
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    m = np.isfinite(x) & np.isfinite(y)
    x = x[m]
    y = y[m]
    if x.size == 0:
        return float("nan")
    bx = np.floor(x / bin_px).astype(int)
    by = np.floor(y / bin_px).astype(int)
    pairs = np.stack([bx, by], axis=1)
    _, counts = np.unique(pairs, axis=0, return_counts=True)
    p = counts / counts.sum()
    return float(-(p * np.log2(p)).sum())


def convex_hull_area_perimeter(points_xy: np.ndarray) -> Tuple[float, float]:
    pts = np.asarray(points_xy, dtype=float)
    pts = pts[np.isfinite(pts).all(axis=1)]
    if pts.shape[0] < 3 or ConvexHull is None:
        return 0.0, 0.0
    try:
        hull = ConvexHull(pts)
        # In 2D: hull.volume == area, hull.area == perimeter
        return float(hull.volume), float(hull.area)
    except Exception:
        return 0.0, 0.0


def box_count_fractal_dimension(
    x: np.ndarray,
    y: np.ndarray,
    box_sizes: Tuple[int, ...] = (4, 8, 16, 32, 64, 128, 256),
) -> float:
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    m = np.isfinite(x) & np.isfinite(y)
    x = x[m]
    y = y[m]
    if x.size < 2:
        return float("nan")

    Ns: List[int] = []
    Es: List[int] = []
    for e in box_sizes:
        bx = np.floor(x / e).astype(int)
        by = np.floor(y / e).astype(int)
        pairs = np.stack([bx, by], axis=1)
        N = np.unique(pairs, axis=0).shape[0]
        if N > 0:
            Ns.append(int(N))
            Es.append(int(e))

    if len(Ns) < 2:
        return float("nan")

    logN = np.log(Ns)
    logInv = np.log([1.0 / e for e in Es])
    slope = float(np.polyfit(logInv, logN, 1)[0])
    return slope


def rqa_revisit_determinism(
    cell_ids: np.ndarray,
    l_min: int = RQA_LMIN,
    max_points: int = MAX_RQA_POINTS,
) -> Tuple[float, float]:
    g = np.asarray(cell_ids)
    if g.size == 0:
        return float("nan"), float("nan")
    if g.size > max_points:
        idx = np.linspace(0, g.size - 1, max_points).astype(int)
        g = g[idx]

    N = int(g.size)
    M = (g[:, None] == g[None, :])
    ones = int(M.sum())
    rr = ones / float(N * N)

    # Determinism: fraction of recurrence points forming diagonal lines (>= l_min)
    num = 0
    for k in range(-(N - 1), N):
        diag = np.diagonal(M, offset=k)
        run = 0
        for v in diag:
            if v:
                run += 1
            else:
                if run >= l_min:
                    num += run
                run = 0
        if run >= l_min:
            num += run

    det = (num / ones) if ones > 0 else float("nan")
    return float(rr), float(det)


def compute_derivatives(v: np.ndarray, dt_s: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    """
    v and dt_s are interval-level arrays length M (between samples).
    Returns:
      a: acceleration (length M-1) computed as dv / dt_{i+1}
      j: jerk (length M-2) computed as da / dt_{i+2}
    """
    v = np.asarray(v, dtype=float)
    dt_s = np.asarray(dt_s, dtype=float)
    if v.size < 2:
        return np.array([], dtype=float), np.array([], dtype=float)

    dv = np.diff(v)
    dt2 = dt_s[1:]
    a = np.full_like(dv, np.nan, dtype=float)
    m = np.isfinite(dv) & np.isfinite(dt2) & (dt2 > 0)
    a[m] = dv[m] / dt2[m]

    if a.size < 2:
        return a[np.isfinite(a)], np.array([], dtype=float)

    da = np.diff(a)
    dt3 = dt_s[2:]
    j = np.full_like(da, np.nan, dtype=float)
    m2 = np.isfinite(da) & np.isfinite(dt3) & (dt3 > 0)
    j[m2] = da[m2] / dt3[m2]
    return a[np.isfinite(a)], j[np.isfinite(j)]


def in_center_region(x: np.ndarray, y: np.ndarray, screen_w: float, screen_h: float) -> np.ndarray:
    cx = screen_w / 2.0
    cy = screen_h / 2.0
    w = CENTER_REGION_PCT * screen_w
    h = CENTER_REGION_PCT * screen_h
    x0 = cx - w / 2.0
    x1 = cx + w / 2.0
    y0 = cy - h / 2.0
    y1 = cy + h / 2.0
    return (x >= x0) & (x <= x1) & (y >= y0) & (y <= y1)


def in_edge_region(x: np.ndarray, y: np.ndarray, screen_w: float, screen_h: float) -> np.ndarray:
    mx = EDGE_MARGIN_PCT * screen_w
    my = EDGE_MARGIN_PCT * screen_h
    return (x <= mx) | (x >= (screen_w - mx)) | (y <= my) | (y >= (screen_h - my))

def make_rect(x_min: float, x_max: float, y_min: float, y_max: float) -> Dict[str, float]:
    return {
        "x_min": float(x_min),
        "x_max": float(x_max),
        "y_min": float(y_min),
        "y_max": float(y_max),
    }

def in_rect_scalar(x: float, y: float, rect: Dict[str, float]) -> bool:
    return (
        x >= rect["x_min"]
        and x < rect["x_max"]
        and y >= rect["y_min"]
        and y < rect["y_max"]
    )

def make_flanker_distractor_rects(
    screen_w: float = FLANKER_SCREEN_WIDTH,
    screen_h: float = FLANKER_SCREEN_HEIGHT,
) -> List[Dict[str, float]]:
    scale_x = float(screen_w) / FLANKER_SCREEN_WIDTH if FLANKER_SCREEN_WIDTH > 0 else 1.0
    scale_y = float(screen_h) / FLANKER_SCREEN_HEIGHT if FLANKER_SCREEN_HEIGHT > 0 else 1.0
    stimulus_left = (
        FLANKER_SCREEN_WIDTH
        - (FLANKER_ARROW_COUNT * FLANKER_ARROW_SIZE + (FLANKER_ARROW_COUNT - 1) * FLANKER_ARROW_GAP)
    ) / 2.0
    stimulus_top = (FLANKER_SCREEN_HEIGHT - FLANKER_ARROW_SIZE) / 2.0

    rects: List[Dict[str, float]] = []
    for i in (0, 1, 3, 4):
        x0 = stimulus_left + i * (FLANKER_ARROW_SIZE + FLANKER_ARROW_GAP)
        y0 = stimulus_top
        rects.append(
            make_rect(
                x0 * scale_x,
                (x0 + FLANKER_ARROW_SIZE) * scale_x,
                y0 * scale_y,
                (y0 + FLANKER_ARROW_SIZE) * scale_y,
            )
        )
    return rects

def count_aoi_transitions(labels: List[str], negative_labels: set[str]) -> int:
    binary: List[str] = []
    for label in labels:
        if label == "aoi":
            binary.append("aoi")
        elif label in negative_labels:
            binary.append("non_aoi")
    if len(binary) < 2:
        return 0
    return int(sum(1 for a, b in zip(binary[:-1], binary[1:]) if a != b))


@dataclass
class Event:
    kind: str              # 'fixation' or 'saccade'
    start_idx: int         # sample index (on-screen filtered series)
    end_idx: int           # sample index inclusive
    start_ms: int
    end_ms: int
    xs: np.ndarray
    ys: np.ndarray


def segment_fix_sac(
    x: np.ndarray,
    y: np.ndarray,
    t_ms: np.ndarray,
    max_gap_ms: float = MAX_GAP_MS,
    min_fix_ms: float = MIN_FIX_MS,
    min_sac_ms: float = MIN_SAC_MS,
    base_thr_px_s: float = BASE_SACCADE_THR_PX_S,
    mad_multiplier: float = MAD_MULT,
) -> Tuple[List[Event], List[Event], Dict[str, Any]]:
    """
    I-VT segmentation on on-screen samples:
      - build interval velocities
      - mark intervals as saccade if speed > threshold
      - group consecutive intervals into events (break on invalid/gap intervals)
    """
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    t_ms = np.asarray(t_ms, dtype=np.int64)

    n = int(x.size)
    if n < 2:
        return [], [], {
            "threshold_px_s": float("nan"),
            "speed": np.array([], dtype=float),
            "vx": np.array([], dtype=float),
            "vy": np.array([], dtype=float),
            "dt_s": np.array([], dtype=float),
            "theta": np.array([], dtype=float),
        }

    dt_ms = np.diff(t_ms).astype(float)  # length M
    valid_interval = (dt_ms > 0) & (dt_ms <= max_gap_ms)
    dt_s = dt_ms / 1000.0

    dx = np.diff(x)
    dy = np.diff(y)
    dist = np.sqrt(dx * dx + dy * dy)

    M = int(dt_ms.size)
    vx = np.full(M, np.nan, dtype=float)
    vy = np.full(M, np.nan, dtype=float)
    speed = np.full(M, np.nan, dtype=float)
    theta = np.full(M, np.nan, dtype=float)

    vx[valid_interval] = dx[valid_interval] / dt_s[valid_interval]
    vy[valid_interval] = dy[valid_interval] / dt_s[valid_interval]
    speed[valid_interval] = dist[valid_interval] / dt_s[valid_interval]
    theta[valid_interval] = np.arctan2(dy[valid_interval], dx[valid_interval])

    sp_valid = speed[np.isfinite(speed)]
    if sp_valid.size >= 10:
        med = float(np.median(sp_valid))
        mad = float(np.median(np.abs(sp_valid - med)))
        thr = max(base_thr_px_s, med + mad_multiplier * mad)
    elif sp_valid.size > 0:
        thr = max(base_thr_px_s, float(np.median(sp_valid)))
    else:
        thr = base_thr_px_s

    is_saccade_interval = valid_interval & (speed > thr)

    # Segment intervals into events
    segments: List[Tuple[str, int, int]] = []
    cur_label: Optional[str] = None
    cur_start: Optional[int] = None
    for i in range(M):
        if not valid_interval[i]:
            if cur_label is not None:
                segments.append((cur_label, int(cur_start), i - 1))  # close
                cur_label, cur_start = None, None
            continue

        label = "saccade" if bool(is_saccade_interval[i]) else "fixation"
        if cur_label is None:
            cur_label, cur_start = label, i
        elif label != cur_label:
            segments.append((cur_label, int(cur_start), i - 1))
            cur_label, cur_start = label, i

    if cur_label is not None:
        segments.append((cur_label, int(cur_start), M - 1))

    fixations: List[Event] = []
    saccades: List[Event] = []
    for label, i0, i1 in segments:
        s0 = i0
        s1 = i1 + 1  # convert interval indices to sample end index
        start = int(t_ms[s0])
        end = int(t_ms[s1])
        dur = float(end - start)
        if label == "fixation" and dur >= min_fix_ms:
            fixations.append(Event("fixation", s0, s1, start, end, x[s0:s1 + 1], y[s0:s1 + 1]))
        elif label == "saccade" and dur >= min_sac_ms:
            saccades.append(Event("saccade", s0, s1, start, end, x[s0:s1 + 1], y[s0:s1 + 1]))

    meta = {"threshold_px_s": float(thr), "speed": speed, "vx": vx, "vy": vy, "dt_s": dt_s, "theta": theta}
    return fixations, saccades, meta


def parse_timestamp_column(ts: pd.Series) -> np.ndarray:
    """Return timestamps in ms as int64."""
    if np.issubdtype(ts.dtype, np.number):
        return ts.astype("int64").to_numpy()
    dt = pd.to_datetime(ts, errors="coerce", utc=True)
    if dt.isna().any():
        raise ValueError("timestamp column contains non-numeric values that could not be parsed as datetime.")
    # ns -> ms
    return (dt.view("int64") // 1_000_000).astype("int64").to_numpy()


def ms_to_iso(ms: int, tz_name: str = "Asia/Seoul") -> str:
    """
    Format like: 2025-09-18T12:26:20.660000+09
    (matches template style: +09 without ':00')
    """
    if ZoneInfo is not None:
        tz = ZoneInfo(tz_name)
        dt = datetime.fromtimestamp(ms / 1000.0, tz=tz)
        s = dt.isoformat(timespec="microseconds")
        return s.replace("+09:00", "+09")
    tz = timezone(timedelta(hours=9))
    dt = datetime.fromtimestamp(ms / 1000.0, tz=tz)
    s = dt.isoformat(timespec="microseconds")
    return s.replace("+09:00", "+09")


def compute_window_base_vectors(
    dfw: pd.DataFrame,
    w_start_ms: int,
    w_end_ms: int,
    screen_w: float,
    screen_h: float,
    context: Optional[str] = None,
    aoi_rect: Optional[Dict[str, float]] = None,
) -> Dict[str, np.ndarray]:
    """
    Compute base vectors (one per PI base name; each is either:
      - a vector of observations, or
      - a length-1 vector for window-level scalar PIs)
    """
    out: Dict[str, np.ndarray] = {}

    if dfw is None or dfw.shape[0] == 0:
        return out

    dfw = dfw.sort_values("timestamp")
    x_all = dfw["x"].to_numpy(dtype=float)
    y_all = dfw["y"].to_numpy(dtype=float)
    t_all = dfw["timestamp"].to_numpy(dtype=np.int64)

    cx = screen_w / 2.0
    cy = screen_h / 2.0

    finite = np.isfinite(x_all) & np.isfinite(y_all) & np.isfinite(t_all.astype(float))
    on_screen = finite & (x_all >= 0) & (x_all <= screen_w) & (y_all >= 0) & (y_all <= screen_h)

    # Region / on-off durations use dwell time on the ORIGINAL timeline
    if t_all.size >= 2:
        dt_ms = np.diff(t_all).astype(float)
        dwell = np.where((dt_ms > 0) & (dt_ms <= MAX_DWELL_MS), dt_ms, 0.0)

        on_dur = float(np.sum(dwell[on_screen[:-1]]))
        off_dur = float(np.sum(dwell[~on_screen[:-1]]))

        x0 = x_all[:-1]
        y0 = y_all[:-1]
        on_mask = on_screen[:-1]

        edge_dur = float(np.sum(dwell[on_mask & in_edge_region(x0, y0, screen_w, screen_h)]))
        center_dur = float(np.sum(dwell[on_mask & in_center_region(x0, y0, screen_w, screen_h)]))
        left_dur = float(np.sum(dwell[on_mask & (x0 < cx)]))
        right_dur = float(np.sum(dwell[on_mask & (x0 >= cx)]))
        top_dur = float(np.sum(dwell[on_mask & (y0 < cy)]))
        bottom_dur = float(np.sum(dwell[on_mask & (y0 >= cy)]))
    else:
        on_dur = off_dur = edge_dur = center_dur = left_dur = right_dur = top_dur = bottom_dur = float("nan")

    window_dur_ms = float(w_end_ms - w_start_ms)
    window_dur_s = window_dur_ms / 1000.0 if window_dur_ms > 0 else float("nan")

    def scalar(v: float) -> np.ndarray:
        return np.array([v], dtype=float)

    out["on_screen_duration"] = scalar(on_dur)
    out["on_screen_gaze_ratio"] = scalar(on_dur / window_dur_ms if (np.isfinite(on_dur) and window_dur_ms > 0) else float("nan"))
    out["off_screen_duration"] = scalar(off_dur)
    out["off_screen_gaze_ratio"] = scalar(off_dur / window_dur_ms if (np.isfinite(off_dur) and window_dur_ms > 0) else float("nan"))
    out["edge_gaze_ratio"] = scalar(edge_dur / window_dur_ms if (np.isfinite(edge_dur) and window_dur_ms > 0) else float("nan"))
    out["left_half_gaze_ratio"] = scalar(left_dur / window_dur_ms if (np.isfinite(left_dur) and window_dur_ms > 0) else float("nan"))
    out["right_half_gaze_ratio"] = scalar(right_dur / window_dur_ms if (np.isfinite(right_dur) and window_dur_ms > 0) else float("nan"))
    out["top_half_gaze_ratio"] = scalar(top_dur / window_dur_ms if (np.isfinite(top_dur) and window_dur_ms > 0) else float("nan"))
    out["bottom_half_gaze_ratio"] = scalar(bottom_dur / window_dur_ms if (np.isfinite(bottom_dur) and window_dur_ms > 0) else float("nan"))
    out["central_gaze_ratio"] = scalar(center_dur / window_dur_ms if (np.isfinite(center_dur) and window_dur_ms > 0) else float("nan"))

    # Most PIs use on-screen samples only
    xv = x_all[on_screen]
    yv = y_all[on_screen]
    tv = t_all[on_screen]
    if xv.size == 0:
        return out

    out["gaze_position_x"] = xv
    out["gaze_position_y"] = yv

    # Interval-level kinematics between on-screen samples
    if tv.size >= 2:
        dt_ms_v = np.diff(tv).astype(float)
        dx = np.diff(xv)
        dy = np.diff(yv)
        dist = np.sqrt(dx * dx + dy * dy)

        valid_int = (dt_ms_v > 0) & (dt_ms_v <= MAX_GAP_MS)
        dt_s_v = dt_ms_v / 1000.0

        M = int(dt_ms_v.size)
        vx = np.full(M, np.nan, dtype=float)
        vy = np.full(M, np.nan, dtype=float)
        speed = np.full(M, np.nan, dtype=float)
        theta = np.full(M, np.nan, dtype=float)

        vx[valid_int] = dx[valid_int] / dt_s_v[valid_int]
        vy[valid_int] = dy[valid_int] / dt_s_v[valid_int]
        speed[valid_int] = dist[valid_int] / dt_s_v[valid_int]
        theta[valid_int] = np.arctan2(dy[valid_int], dx[valid_int])
    else:
        vx = vy = speed = theta = np.array([], dtype=float)
        dist = np.array([], dtype=float)
        dt_s_v = np.array([], dtype=float)
        valid_int = np.array([], dtype=bool)

    out["gaze_velocity"] = speed[np.isfinite(speed)]
    out["gaze_velocity_x"] = vx[np.isfinite(vx)]
    out["gaze_velocity_y"] = vy[np.isfinite(vy)]

    # Derivatives
    a, j = compute_derivatives(speed, dt_s_v) if speed.size else (np.array([], dtype=float), np.array([], dtype=float))
    ax, jx = compute_derivatives(vx, dt_s_v) if vx.size else (np.array([], dtype=float), np.array([], dtype=float))
    ay, jy = compute_derivatives(vy, dt_s_v) if vy.size else (np.array([], dtype=float), np.array([], dtype=float))

    out["gaze_acceleration"] = a
    out["gaze_acceleration_x"] = ax
    out["gaze_acceleration_y"] = ay
    out["gaze_jerk"] = j
    out["gaze_jerk_x"] = jx
    out["gaze_jerk_y"] = jy

    # Gaze path metrics (scalar per window)
    path_length = float(np.nansum(dist[valid_int])) if dist.size else 0.0
    out["gaze_path_length"] = np.array([path_length], dtype=float)

    disp = float(np.sqrt((xv[-1] - xv[0]) ** 2 + (yv[-1] - yv[0]) ** 2)) if xv.size >= 2 else 0.0
    out["gaze_path_displacement"] = np.array([disp], dtype=float)
    out["gaze_path_curvature"] = np.array([path_length / disp if disp > 0 else float("nan")], dtype=float)

    # Bbox + dispersion (scalar per window)
    xr = float(np.nanmax(xv) - np.nanmin(xv))
    yr = float(np.nanmax(yv) - np.nanmin(yv))
    out["gaze_bbox_width"] = np.array([xr], dtype=float)
    out["gaze_bbox_height"] = np.array([yr], dtype=float)
    out["gaze_bbox_aspect_ratio"] = np.array([xr / yr if yr != 0 else float("nan")], dtype=float)
    out["range_based_gaze_dispersion"] = np.array([xr + yr], dtype=float)
    out["rms_based_gaze_dispersion"] = np.array([float(np.sqrt(np.nanvar(xv, ddof=1) + np.nanvar(yv, ddof=1))) if xv.size > 1 else 0.0], dtype=float)
    out["bbox_area_based_gaze_dispersion"] = np.array([xr * yr], dtype=float)
    out["gaze_covariance"] = np.array([float(np.cov(xv, yv, ddof=1)[0, 1]) if xv.size > 1 else 0.0], dtype=float)
    out["gaze_entropy"] = np.array([entropy_2d(xv, yv)], dtype=float)

    d_center = np.sqrt((xv - cx) ** 2 + (yv - cy) ** 2)
    out["gaze_offset_to_screen_center"] = d_center
    dmax = float(np.nanmax(d_center)) if d_center.size else float("nan")
    dmean = float(np.nanmean(d_center)) if d_center.size else float("nan")
    out["central_bias"] = np.array([1.0 - dmean / dmax if (np.isfinite(dmean) and np.isfinite(dmax) and dmax > 0) else float("nan")], dtype=float)

    # Fixation/saccade segmentation on on-screen samples
    fixations, saccades, meta = segment_fix_sac(xv, yv, tv)
    out["fixation_rate"] = np.array([len(fixations) / window_dur_s if window_dur_s > 0 else float("nan")], dtype=float)
    out["saccade_rate"] = np.array([len(saccades) / window_dur_s if window_dur_s > 0 else float("nan")], dtype=float)
    out["fixation_saccade_ratio"] = np.array([len(fixations) / len(saccades) if len(saccades) > 0 else float("nan")], dtype=float)

    # Fixation metrics (event-level vectors)
    fix_durs: List[float] = []
    fix_bbox_w: List[float] = []
    fix_bbox_h: List[float] = []
    fix_bbox_ar: List[float] = []
    fix_disp_range: List[float] = []
    fix_disp_rms: List[float] = []
    fix_disp_area: List[float] = []
    fix_ent: List[float] = []
    fix_centroids: List[Tuple[float, float]] = []
    drift_v: List[float] = []
    drift_vx: List[float] = []
    drift_vy: List[float] = []

    for ev in fixations:
        fix_durs.append(float(ev.end_ms - ev.start_ms))
        w = float(np.nanmax(ev.xs) - np.nanmin(ev.xs))
        h = float(np.nanmax(ev.ys) - np.nanmin(ev.ys))
        fix_bbox_w.append(w)
        fix_bbox_h.append(h)
        fix_bbox_ar.append(w / h if h != 0 else float("nan"))
        fix_disp_range.append(w + h)
        fix_disp_rms.append(float(np.sqrt(np.nanvar(ev.xs, ddof=1) + np.nanvar(ev.ys, ddof=1))) if ev.xs.size > 1 else 0.0)
        fix_disp_area.append(w * h)
        fix_ent.append(entropy_2d(ev.xs, ev.ys))
        fix_centroids.append((float(np.nanmean(ev.xs)), float(np.nanmean(ev.ys))))

        i0 = ev.start_idx
        i1 = ev.end_idx - 1
        sp = meta["speed"][i0:i1 + 1]
        vx_ev = meta["vx"][i0:i1 + 1]
        vy_ev = meta["vy"][i0:i1 + 1]
        drift_v.extend(sp[np.isfinite(sp)].tolist())
        drift_vx.extend(vx_ev[np.isfinite(vx_ev)].tolist())
        drift_vy.extend(vy_ev[np.isfinite(vy_ev)].tolist())

    if aoi_rect is not None:
        task_key = str(context or "").lower()
        flanker_distractors = make_flanker_distractor_rects(screen_w, screen_h) if task_key == "flanker" else []
        aoi_dur = 0.0
        distractor_dur = 0.0
        empty_dur = 0.0
        aoi_fix_durs: List[float] = []
        aoi_first_latency = float("nan")
        aoi_first_duration = float("nan")
        fixation_offsets_to_aoi: List[float] = []
        aoi_labels: List[str] = []
        aoi_center_x = (aoi_rect["x_min"] + aoi_rect["x_max"]) / 2.0
        aoi_center_y = (aoi_rect["y_min"] + aoi_rect["y_max"]) / 2.0

        for ev in fixations:
            ev_cx = float(np.nanmean(ev.xs))
            ev_cy = float(np.nanmean(ev.ys))
            ev_dur = float(ev.end_ms - ev.start_ms)
            fixation_offsets_to_aoi.append(float(np.sqrt((ev_cx - aoi_center_x) ** 2 + (ev_cy - aoi_center_y) ** 2)))

            if in_rect_scalar(ev_cx, ev_cy, aoi_rect):
                label = "aoi"
                aoi_dur += ev_dur
                aoi_fix_durs.append(ev_dur)
                if not np.isfinite(aoi_first_latency):
                    aoi_first_latency = float(ev.start_ms - w_start_ms)
                    aoi_first_duration = ev_dur
            elif task_key == "flanker":
                is_distractor = any(in_rect_scalar(ev_cx, ev_cy, rect) for rect in flanker_distractors)
                if is_distractor:
                    label = "distractor"
                    distractor_dur += ev_dur
                else:
                    label = "empty"
                    empty_dur += ev_dur
            else:
                label = "distractor"
                distractor_dur += ev_dur

            aoi_labels.append(label)

        if task_key == "flanker":
            non_aoi_dur = distractor_dur
            transition_count = float(count_aoi_transitions(aoi_labels, {"distractor"}))
        else:
            non_aoi_dur = distractor_dur
            transition_count = float(count_aoi_transitions(aoi_labels, {"distractor"}))

        aoi_dwell_time = float("nan")
        if t_all.size >= 2:
            dt_ms = np.diff(t_all).astype(float)
            dwell = np.where((dt_ms > 0) & (dt_ms <= MAX_DWELL_MS), dt_ms, 0.0)
            x0 = x_all[:-1]
            y0 = y_all[:-1]
            on_mask = on_screen[:-1]
            aoi_mask = (
                on_mask
                & (x0 >= aoi_rect["x_min"])
                & (x0 < aoi_rect["x_max"])
                & (y0 >= aoi_rect["y_min"])
                & (y0 < aoi_rect["y_max"])
            )
            aoi_dwell_time = float(np.sum(dwell[aoi_mask]))

        binary_labels = ["aoi" if label == "aoi" else "non_aoi" for label in aoi_labels]
        transition_matrix = np.zeros((2, 2), dtype=float)
        label_to_idx = {"aoi": 0, "non_aoi": 1}
        for prev_label, next_label in zip(binary_labels[:-1], binary_labels[1:]):
            transition_matrix[label_to_idx[prev_label], label_to_idx[next_label]] += 1.0
        transition_total = float(np.sum(transition_matrix))
        if transition_total > 0:
            probs = transition_matrix.ravel() / transition_total
            probs = probs[probs > 0]
            matrix_entropy = float(-np.sum(probs * np.log2(probs)))
            matrix_sparsity = float(np.count_nonzero(transition_matrix == 0.0) / transition_matrix.size)
        else:
            matrix_entropy = 0.0
            matrix_sparsity = 1.0

        aoi_fix_count = len(aoi_fix_durs)
        aoi_transition_rate = transition_count / window_dur_s if window_dur_s > 0 else float("nan")
        aoi_revisit_rate = max(aoi_fix_count - 1, 0) / window_dur_s if window_dur_s > 0 else float("nan")
        aoi_fixation_rate = aoi_fix_count / window_dur_s if window_dur_s > 0 else float("nan")
        sequence_distance = float(sum(1 for label in binary_labels if label != "aoi"))

        out["aoi_fixation_duration"] = np.array([aoi_dur], dtype=float)
        out["aoi_dwell_time"] = np.array([aoi_dwell_time], dtype=float)
        out["aoi_revisit_rate"] = np.array([aoi_revisit_rate], dtype=float)
        out["aoi_transition_rate"] = np.array([aoi_transition_rate], dtype=float)
        out["aoi_first_fixation_latency"] = np.array([aoi_first_latency], dtype=float)
        out["aoi_first_fixation_duration"] = np.array([aoi_first_duration], dtype=float)
        out["aoi_fixation_rate"] = np.array([aoi_fixation_rate], dtype=float)
        out["aoi_transition_matrix_entropy"] = np.array([matrix_entropy], dtype=float)
        out["aoi_transition_matrix_sparsity"] = np.array([matrix_sparsity], dtype=float)
        out["fixation_offset_to_aoi"] = np.asarray(fixation_offsets_to_aoi, dtype=float)
        out["aoi_sequence_levenshtein_distance"] = np.array([sequence_distance], dtype=float)
        out["distractor_fixation_duration"] = np.array([distractor_dur], dtype=float)
        out["empty_space_fixation_duration"] = np.array([empty_dur], dtype=float)
        out["non_aoi_fixation_duration"] = np.array([non_aoi_dur], dtype=float)
        out["aoi_non_aoi_transition_count"] = np.array([transition_count], dtype=float)

    out["fixation_duration"] = np.asarray(fix_durs, dtype=float)
    out["fixation_bbox_width"] = np.asarray(fix_bbox_w, dtype=float)
    out["fixation_bbox_height"] = np.asarray(fix_bbox_h, dtype=float)
    out["fixation_bbox_aspect_ratio"] = np.asarray(fix_bbox_ar, dtype=float)
    out["range_based_fixation_dispersion"] = np.asarray(fix_disp_range, dtype=float)
    out["rms_based_fixation_dispersion"] = np.asarray(fix_disp_rms, dtype=float)
    out["bbox_area_based_fixation_dispersion"] = np.asarray(fix_disp_area, dtype=float)
    out["fixation_entropy"] = np.asarray(fix_ent, dtype=float)
    out["fixational_drift_velocity"] = np.asarray(drift_v, dtype=float)
    out["fixational_drift_velocity_x"] = np.asarray(drift_vx, dtype=float)
    out["fixational_drift_velocity_y"] = np.asarray(drift_vy, dtype=float)

    if fixations:
        first = fixations[0]
        out["first_fixation_latency"] = np.array([float(first.start_ms - w_start_ms)], dtype=float)
        out["first_fixation_duration"] = np.array([float(first.end_ms - first.start_ms)], dtype=float)
    else:
        out["first_fixation_latency"] = np.array([float("nan")], dtype=float)
        out["first_fixation_duration"] = np.array([float("nan")], dtype=float)

    # Fixation centroid dispersion + covariance + convex hull (scalar per window)
    if fix_centroids:
        cxs = np.asarray([p[0] for p in fix_centroids], dtype=float)
        cys = np.asarray([p[1] for p in fix_centroids], dtype=float)
        cxr = float(np.nanmax(cxs) - np.nanmin(cxs))
        cyr = float(np.nanmax(cys) - np.nanmin(cys))
        out["range_based_fixation_centroid_dispersion"] = np.array([cxr + cyr], dtype=float)
        out["rms_based_fixation_centroid_dispersion"] = np.array([float(np.sqrt(np.nanvar(cxs, ddof=1) + np.nanvar(cys, ddof=1))) if cxs.size > 1 else 0.0], dtype=float)
        out["bbox_area_based_fixation_centroid_dispersion"] = np.array([cxr * cyr], dtype=float)
        out["fixation_centroid_covariance"] = np.array([float(np.cov(cxs, cys, ddof=1)[0, 1]) if cxs.size > 1 else 0.0], dtype=float)
        area, per = convex_hull_area_perimeter(np.stack([cxs, cys], axis=1))
        out["fixation_convex_hull_area"] = np.array([area], dtype=float)
        out["fixation_convex_hull_perimeter"] = np.array([per], dtype=float)
    else:
        out["range_based_fixation_centroid_dispersion"] = np.array([float("nan")], dtype=float)
        out["rms_based_fixation_centroid_dispersion"] = np.array([float("nan")], dtype=float)
        out["bbox_area_based_fixation_centroid_dispersion"] = np.array([float("nan")], dtype=float)
        out["fixation_centroid_covariance"] = np.array([float("nan")], dtype=float)
        out["fixation_convex_hull_area"] = np.array([float("nan")], dtype=float)
        out["fixation_convex_hull_perimeter"] = np.array([float("nan")], dtype=float)

    # Inter-fixation interval
    if len(fixations) >= 2:
        ifi = [float(fixations[i + 1].start_ms - fixations[i].end_ms) for i in range(len(fixations) - 1)]
        out["inter_fixation_interval"] = np.asarray(ifi, dtype=float)
    else:
        out["inter_fixation_interval"] = np.array([], dtype=float)

    # Saccade metrics (event-level vectors)
    sac_durs: List[float] = []
    sac_bbox_w: List[float] = []
    sac_bbox_h: List[float] = []
    sac_bbox_ar: List[float] = []
    sac_disp_range: List[float] = []
    sac_disp_rms: List[float] = []
    sac_disp_area: List[float] = []
    sac_ent: List[float] = []
    sac_displacement: List[float] = []
    sac_dir: List[float] = []
    sac_path_len: List[float] = []
    sac_path_curv: List[float] = []
    sac_v: List[float] = []
    sac_vx: List[float] = []
    sac_vy: List[float] = []
    sac_a: List[float] = []
    sac_ax: List[float] = []
    sac_ay: List[float] = []
    sac_j: List[float] = []
    sac_jx: List[float] = []
    sac_jy: List[float] = []
    sac_angvel: List[float] = []
    endpoints: List[Tuple[float, float]] = []

    for ev in saccades:
        sac_durs.append(float(ev.end_ms - ev.start_ms))
        w = float(np.nanmax(ev.xs) - np.nanmin(ev.xs))
        h = float(np.nanmax(ev.ys) - np.nanmin(ev.ys))
        sac_bbox_w.append(w)
        sac_bbox_h.append(h)
        sac_bbox_ar.append(w / h if h != 0 else float("nan"))
        sac_disp_range.append(w + h)
        sac_disp_rms.append(float(np.sqrt(np.nanvar(ev.xs, ddof=1) + np.nanvar(ev.ys, ddof=1))) if ev.xs.size > 1 else 0.0)
        sac_disp_area.append(w * h)
        sac_ent.append(entropy_2d(ev.xs, ev.ys))

        dx_e = float(ev.xs[-1] - ev.xs[0])
        dy_e = float(ev.ys[-1] - ev.ys[0])
        disp_e = float(np.sqrt(dx_e * dx_e + dy_e * dy_e))
        sac_displacement.append(disp_e)
        sac_dir.append(float(np.arctan2(dy_e, dx_e)))

        if ev.xs.size >= 2:
            dxe = np.diff(ev.xs)
            dye = np.diff(ev.ys)
            sac_path_len.append(float(np.nansum(np.sqrt(dxe * dxe + dye * dye))))
        else:
            sac_path_len.append(0.0)
        sac_path_curv.append(sac_path_len[-1] / disp_e if disp_e > 0 else float("nan"))

        endpoints.append((float(ev.xs[-1]), float(ev.ys[-1])))

        i0 = ev.start_idx
        i1 = ev.end_idx - 1
        sp = meta["speed"][i0:i1 + 1]
        vx_ev = meta["vx"][i0:i1 + 1]
        vy_ev = meta["vy"][i0:i1 + 1]
        dt_ev = meta["dt_s"][i0:i1 + 1]

        sac_v.extend(sp[np.isfinite(sp)].tolist())
        sac_vx.extend(vx_ev[np.isfinite(vx_ev)].tolist())
        sac_vy.extend(vy_ev[np.isfinite(vy_ev)].tolist())

        acc, jer = compute_derivatives(sp, dt_ev)
        accx, jerx = compute_derivatives(vx_ev, dt_ev)
        accy, jery = compute_derivatives(vy_ev, dt_ev)
        sac_a.extend(acc.tolist())
        sac_j.extend(jer.tolist())
        sac_ax.extend(accx.tolist())
        sac_jx.extend(jerx.tolist())
        sac_ay.extend(accy.tolist())
        sac_jy.extend(jery.tolist())

        # Angular velocity
        th = meta["theta"][i0:i1 + 1]
        m_th = np.isfinite(th)
        if np.count_nonzero(m_th) >= 2:
            th2 = th[m_th]
            th_un = np.unwrap(th2)
            dt_th = dt_ev[m_th][1:]
            dth = np.diff(th_un)
            m_ok = np.isfinite(dt_th) & (dt_th > 0)
            sac_angvel.extend((dth[m_ok] / dt_th[m_ok]).tolist())

    out["saccade_duration"] = np.asarray(sac_durs, dtype=float)
    out["saccade_bbox_width"] = np.asarray(sac_bbox_w, dtype=float)
    out["saccade_bbox_height"] = np.asarray(sac_bbox_h, dtype=float)
    out["saccade_bbox_aspect_ratio"] = np.asarray(sac_bbox_ar, dtype=float)
    out["range_based_saccade_dispersion"] = np.asarray(sac_disp_range, dtype=float)
    out["rms_based_saccade_dispersion"] = np.asarray(sac_disp_rms, dtype=float)
    out["bbox_area_based_saccade_dispersion"] = np.asarray(sac_disp_area, dtype=float)
    out["saccade_entropy"] = np.asarray(sac_ent, dtype=float)
    out["saccade_displacement"] = np.asarray(sac_displacement, dtype=float)
    out["saccade_velocity"] = np.asarray(sac_v, dtype=float)
    out["saccade_velocity_x"] = np.asarray(sac_vx, dtype=float)
    out["saccade_velocity_y"] = np.asarray(sac_vy, dtype=float)
    out["saccade_acceleration"] = np.asarray(sac_a, dtype=float)
    out["saccade_acceleration_x"] = np.asarray(sac_ax, dtype=float)
    out["saccade_acceleration_y"] = np.asarray(sac_ay, dtype=float)
    out["saccade_jerk"] = np.asarray(sac_j, dtype=float)
    out["saccade_jerk_x"] = np.asarray(sac_jx, dtype=float)
    out["saccade_jerk_y"] = np.asarray(sac_jy, dtype=float)
    out["saccade_direction"] = np.asarray(sac_dir, dtype=float)
    out["saccade_angular_velocity"] = np.asarray(sac_angvel, dtype=float)
    out["saccade_path_length"] = np.asarray(sac_path_len, dtype=float)
    out["saccade_path_curvature"] = np.asarray(sac_path_curv, dtype=float)

    if saccades:
        first = saccades[0]
        out["first_saccade_latency"] = np.array([float(first.start_ms - w_start_ms)], dtype=float)
        out["first_saccade_duration"] = np.array([float(first.end_ms - first.start_ms)], dtype=float)
    else:
        out["first_saccade_latency"] = np.array([float("nan")], dtype=float)
        out["first_saccade_duration"] = np.array([float("nan")], dtype=float)

    if len(saccades) >= 2:
        isi = [float(saccades[i + 1].start_ms - saccades[i].end_ms) for i in range(len(saccades) - 1)]
        out["inter_saccade_interval"] = np.asarray(isi, dtype=float)
    else:
        out["inter_saccade_interval"] = np.array([], dtype=float)

    # Horizontal/vertical saccade rates (scalar per window)
    if sac_dir:
        dirs = np.asarray(sac_dir, dtype=float)
        hmask = (angular_diff(dirs, 0.0) < SACCADE_AXIS_THR_RAD) | (angular_diff(dirs, np.pi) < SACCADE_AXIS_THR_RAD)
        vmask = (angular_diff(dirs, np.pi / 2) < SACCADE_AXIS_THR_RAD) | (angular_diff(dirs, -np.pi / 2) < SACCADE_AXIS_THR_RAD)
        out["horizontal_saccade_rate"] = np.array([int(hmask.sum()) / window_dur_s if window_dur_s > 0 else float("nan")], dtype=float)
        out["vertical_saccade_rate"] = np.array([int(vmask.sum()) / window_dur_s if window_dur_s > 0 else float("nan")], dtype=float)
    else:
        out["horizontal_saccade_rate"] = np.array([float("nan")], dtype=float)
        out["vertical_saccade_rate"] = np.array([float("nan")], dtype=float)

    # Endpoint dispersion + covariance (scalar per window)
    if endpoints:
        ex = np.asarray([p[0] for p in endpoints], dtype=float)
        ey = np.asarray([p[1] for p in endpoints], dtype=float)
        exr = float(np.nanmax(ex) - np.nanmin(ex))
        eyr = float(np.nanmax(ey) - np.nanmin(ey))
        out["range_based_saccade_endpoint_dispersion"] = np.array([exr + eyr], dtype=float)
        out["rms_based_saccade_endpoint_dispersion"] = np.array([float(np.sqrt(np.nanvar(ex, ddof=1) + np.nanvar(ey, ddof=1))) if ex.size > 1 else 0.0], dtype=float)
        out["bbox_area_based_saccade_endpoint_dispersion"] = np.array([exr * eyr], dtype=float)
        out["saccade_endpoint_covariance"] = np.array([float(np.cov(ex, ey, ddof=1)[0, 1]) if ex.size > 1 else 0.0], dtype=float)
    else:
        out["range_based_saccade_endpoint_dispersion"] = np.array([float("nan")], dtype=float)
        out["rms_based_saccade_endpoint_dispersion"] = np.array([float("nan")], dtype=float)
        out["bbox_area_based_saccade_endpoint_dispersion"] = np.array([float("nan")], dtype=float)
        out["saccade_endpoint_covariance"] = np.array([float("nan")], dtype=float)

    # Scanpath metrics (scalar per window)
    out["scanpath_fractal_dimension"] = np.array([box_count_fractal_dimension(xv, yv)], dtype=float)

    bx = np.floor(xv / RQA_BIN_PX).astype(int)
    by = np.floor(yv / RQA_BIN_PX).astype(int)
    cell_ids = bx * 100000 + by
    rr, det = rqa_revisit_determinism(cell_ids)
    out["scanpath_revisit_ratio"] = np.array([rr], dtype=float)
    out["scanpath_determinism_ratio"] = np.array([det], dtype=float)

    # Backtrack saccades (scalar per window)
    if len(sac_dir) >= 2:
        dirs = np.asarray(sac_dir, dtype=float)
        diffs = angular_diff(dirs[1:], dirs[:-1])
        back = np.abs(diffs - np.pi) < BACKTRACK_TAU_RAD
        n_back = int(back.sum())
        total_sac_dur_s = float(np.sum(np.asarray(sac_durs, dtype=float)) / 1000.0) if sac_durs else 0.0
        out["backtrack_saccade_ratio"] = np.array([n_back / len(sac_dir) if len(sac_dir) > 0 else float("nan")], dtype=float)
        out["backtrack_saccade_rate"] = np.array([n_back / total_sac_dur_s if total_sac_dur_s > 0 else float("nan")], dtype=float)
    else:
        out["backtrack_saccade_ratio"] = np.array([float("nan")], dtype=float)
        out["backtrack_saccade_rate"] = np.array([float("nan")], dtype=float)

    return out


def build_pi_from_fullnames(base_vecs: Dict[str, np.ndarray], fullnames: List[str]) -> Dict[str, Any]:
    """
    - fullnames: 엑셀 Name 컬럼 그대로 (예: et_range-based_gaze_dispersion_mean)
    - base_vecs: 기존 compute_window_base_vectors()가 만들어주는 base key 벡터들 (예: range_based_gaze_dispersion)
    - 출력 key는 fullnames 그대로 유지
    """
    pi: Dict[str, Any] = {}

    # base별로 stats를 1번만 계산해서 캐시
    cache: Dict[str, Dict[str, Optional[float]]] = {}

    for fullname in fullnames:
        base_full, stat = split_fullname(fullname)  # (et_xxx, mean)

        # 1) et_ 접두어 제거해서 기존 base_vecs 키로 맞춤
        base = base_full
        if base.startswith("et_"):
            base = base[3:]

        # 2) 엑셀의 base에 하이픈이 있을 수 있으니 내부 키와 맞추기 위해 '_'로 통일
        #    (기존 코드의 내부 키는 range_based..., left_half... 처럼 '_' 기반)
        base_internal = base.replace("-", "_")

        # 3) 원형 통계가 필요한 경우(기존 로직 유지)
        is_circular = (base_internal == "saccade_direction")

        if base_internal not in cache:
            vec = base_vecs.get(base_internal, np.array([], dtype=float))
            cache[base_internal] = stats_from_vec(vec, circular=is_circular)

        v = cache[base_internal].get(stat, None)
        if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
            pi[fullname] = None
        else:
            pi[fullname] = float(v)

    return pi

def build_pi_from_base_vectors(base_vecs: Dict[str, np.ndarray]) -> Dict[str, Any]:
    pi: Dict[str, Any] = {}
    for base_internal in sorted(base_vecs):
        vec = base_vecs.get(base_internal, np.array([], dtype=float))
        is_circular = (base_internal == "saccade_direction")
        stats = stats_from_vec(vec, circular=is_circular)
        for stat in STAT_SUFFIXES:
            key = f"et_{base_internal}_{stat}"
            v = stats.get(stat)
            if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
                pi[key] = None
            else:
                pi[key] = float(v)
    return pi

def build_direct_aoi_pi(base_vecs: Dict[str, np.ndarray]) -> Dict[str, Any]:
    key_map = {
        "distractor_fixation_duration": "et_distractor_fixation_duration",
        "empty_space_fixation_duration": "et_empty_space_fixation_duration",
        "non_aoi_fixation_duration": "et_non_aoi_fixation_duration",
        "aoi_non_aoi_transition_count": "et_aoi_non_aoi_transition_count",
    }
    out: Dict[str, Any] = {}
    for base_key, pi_key in key_map.items():
        if base_key not in base_vecs:
            continue
        vec = np.asarray(base_vecs.get(base_key, np.array([], dtype=float)), dtype=float)
        if vec.size == 0 or not np.isfinite(vec[0]):
            out[pi_key] = None
        else:
            out[pi_key] = float(vec[0])
    return out

def parse_aoi_tasks(value: Any) -> set[str]:
    if value is None:
        return set(DEFAULT_AOI_TASKS)
    if isinstance(value, str):
        items = re.split(r"[,\s]+", value)
    else:
        items = []
        for v in value:
            items.extend(re.split(r"[,\s]+", str(v)))
    return {str(x).strip().lower() for x in items if str(x).strip()}

def aoi_rect_for_task(task: str, args: Any) -> Optional[Dict[str, float]]:
    task_key = str(task or "").lower()
    if task_key not in parse_aoi_tasks(getattr(args, "aoi_tasks", None)):
        return None
    return make_rect(
        float(args.aoi_x_min),
        float(args.aoi_x_max),
        float(args.aoi_y_min),
        float(args.aoi_y_max),
    )



def iter_windows_partial(t_min_ms: int, t_max_ms: int, window_ms: int, stride_ms: int) -> List[Tuple[int, int, int]]:
    """
    stride로 start를 이동시키면서,
    end = min(start + window_ms, t_max_ms) 로 partial window를 허용.

    단, 여기서는 "구간에 샘플이 있는지"는 체크하지 않는다.
    (샘플 체크는 main/process_one_file에서 dfw로 확인하는 게 정확함)
    """
    windows: List[Tuple[int, int, int]] = []
    w = 1
    start = int(t_min_ms)

    # start가 t_max보다 작기만 하면 윈도우 후보 생성
    while start < t_max_ms:
        end = start + window_ms
        if end > t_max_ms:
            break
        if end > start:  # 안전장치
            windows.append((w, start, end))
            w += 1
        start += stride_ms

    return windows


def process_one_file(
    input_csv: str,
    output_json: str,
    client_id: str,
    context: str,
    sensor_modality: str,
    timezone_name: str,
    screen_w: float,
    screen_h: float,
    window_size_s: float,
    stride_s: float,
    task_start_ms: Optional[int] = None,
    task_end_ms: Optional[int] = None,
    pi_excel: Optional[str] = None,
    pi_sheet: str = "eye-tracking",
    aoi_rect: Optional[Dict[str, float]] = None,
) -> None:
    
    try:
        df = pd.read_csv(input_csv, usecols=["timestamp", "x", "y"])
    except EmptyDataError:
        print(f"[SKIP] Empty CSV (no rows): {input_csv}")
        return
    except Exception as e:
        # Some files trigger a pandas parser bug in this environment.
        # Fall back to the stdlib csv reader for the required columns only.
        try:
            rows = []
            with open(input_csv, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(
                        {
                            "timestamp": row.get("timestamp"),
                            "x": row.get("x"),
                            "y": row.get("y"),
                        }
                    )
            df = pd.DataFrame(rows, columns=["timestamp", "x", "y"])
        except Exception as fallback_e:
            raise RuntimeError(f"Failed to read CSV: {input_csv} (pandas={e}, fallback={fallback_e})") from fallback_e
    # 헤더는 있는데 데이터 행이 0개인 케이스
    if df.empty:
        print(f"[SKIP] CSV has header only (0 rows): {input_csv} (cols={list(df.columns)})")
        return
    
    for col in ("timestamp", "x", "y"):
        if col not in df.columns:
            raise ValueError(f"[{input_csv}] Missing required column: {col}")

    # Normalize timestamp column to ms int
    df["timestamp"] = parse_timestamp_column(df["timestamp"])
    df = df.sort_values("timestamp").reset_index(drop=True)

    pi_fullnames = load_pi_fullnames_from_excel(pi_excel, sheet_name=pi_sheet) if pi_excel else None
    
    window_ms = int(round(window_size_s * 1000.0))
    stride_ms = int(round(stride_s * 1000.0))

    # Windowing anchor:
    #   - if task_start_ms/task_end_ms are provided, windows are generated from [start_t, end_t]
    #   - otherwise, fall back to data-driven [min_ts, max_ts]
    if task_start_ms is not None and task_end_ms is not None and int(task_end_ms) > int(task_start_ms):
        t_min = int(task_start_ms)
        t_max = int(task_end_ms)
    else:
        t_min = int(df["timestamp"].min())
        t_max = int(df["timestamp"].max())

    windows = iter_windows_partial(t_min, t_max, window_ms, stride_ms)

    out_obj: Dict[str, Any] = {
        "client_id": client_id,
        "context": context,
        "sensor_modality": sensor_modality,
        "screen": {"width": screen_w, "height": screen_h},
        "config": {"window_size": window_size_s, "stride": stride_s},
        "windows": [],
    }

    for wnum, w_start, w_end in windows:
        dfw = df[(df["timestamp"] >= w_start) & (df["timestamp"] < w_end)]
        base_vecs = compute_window_base_vectors(
            dfw=dfw,
            w_start_ms=w_start,
            w_end_ms=w_end,
            screen_w=screen_w,
            screen_h=screen_h,
            context=context,
            aoi_rect=aoi_rect,
        )

        pi = (
            build_pi_from_fullnames(base_vecs, pi_fullnames)
            if pi_fullnames is not None
            else build_pi_from_base_vectors(base_vecs)
        )
        pi.update(build_direct_aoi_pi(base_vecs))

        out_obj["windows"].append(
            {
                "window_number": int(wnum),
                "start_timestamp": ms_to_iso(int(w_start), tz_name=timezone_name),
                "end_timestamp": ms_to_iso(int(w_end), tz_name=timezone_name),
                "pi": pi,
            }
        )

    out_dir = os.path.dirname(output_json)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
        
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(out_obj, f, ensure_ascii=False, indent=2)

    print(f"[OK] {client_id}/{context} -> {output_json}  (windows={len(out_obj['windows'])})")



def run_batch_mode(args) -> None:
    """
    base_path 아래를 순회하면서
    {base_path}/{date}/{client_id}/{task}/eye-tracking/{archive_id}/eye-tracking.csv
    를 찾아서, 각각을
    {output_base_path}/{client_id}/{task}/eye-tracking_pi.json
    으로 저장.
    """
    base_path = os.path.abspath(args.base_path)
    out_base = os.path.abspath(args.output_base_path)
    client_filter = set(args.only_client_id) if args.only_client_id else None

    task_time_index = None
    if getattr(args, "task_time_json", None):
        try:
            task_time_index = load_task_time_index(args.task_time_json)
            print(f"[INFO] Loaded task_time_json: {args.task_time_json} (pairs={len(task_time_index)})")
        except Exception as e:
            print(f"[WARN] Failed to load task_time_json: {args.task_time_json} ({e}). Falling back to data-driven windowing.")
            task_time_index = None


    n_files = 0

    for root, dirs, files in os.walk(base_path):
        if "eye-tracking.csv" not in files:
            continue

        csv_path = os.path.join(root, "eye-tracking.csv")
        # base_path 기준 상대 경로
        rel = os.path.relpath(csv_path, base_path)
        parts = rel.split(os.sep)

        # 기대 구조: date/client_id/task/eye-tracking/archive_id/eye-tracking.csv
        # parts = [date, client_id, task, "eye-tracking", archive_id, "eye-tracking.csv"]
        if len(parts) < 6:
            print(f"[WARN] Unexpected path structure, skip: {csv_path}")
            continue

        date_str = parts[0]
        client_id = parts[1]
        task = parts[2]
        modality_dir = parts[3]

        if modality_dir != "eye-tracking":
            print(f"[WARN] modality dir is not 'eye-tracking' (got '{modality_dir}'), skip: {csv_path}")
            continue
        
        if task not in TARGET_TASK:
            continue
        
        
        
        window_size_sec = resolve_window_size_for_task(task, args.window_size_sec)
        stride_sec = resolve_stride_for_task(task, args.stride_sec)

        # ✅ 파일명에 반영
        ws_tag = fmt_sec(window_size_sec)
        st_tag = fmt_sec(stride_sec)
        
        

        # 클라이언트 필터링 (테스트용)
        if client_filter is not None and client_id not in client_filter:
            # print(f"[SKIP] client_id {client_id} not in filter")
            continue

        # 출력 경로: {output_base_path}/{client_id}/{task}/eye-tracking_pi.json
        out_dir = os.path.join(out_base, client_id, task)
        output_json = os.path.join(out_dir, f"eye-tracking_pi_{ws_tag}_{st_tag}.json")

        print(
            f"[INFO] Processing: csv={csv_path} -> json={output_json} "
            f"(client={client_id}, task={task}, date={date_str}, window={window_size_sec}, stride={stride_sec})"
        )

        task_start_ms = None
        task_end_ms = None
        if task_time_index is not None:
            key = (str(client_id), str(task))
            if key in task_time_index:
                task_start_ms, task_end_ms = task_time_index[key]

        process_one_file(

            input_csv=csv_path,
            output_json=output_json,
            client_id=client_id,
            context=task,  # task를 곧바로 context로 사용
            sensor_modality="eye-tracking",
            timezone_name=args.timezone,
            screen_w=float(args.screen_width_px),
            screen_h=float(args.screen_height_px),
            window_size_s=window_size_sec,
            stride_s=stride_sec,
            task_start_ms=task_start_ms,
            task_end_ms=task_end_ms,
            pi_excel=args.pi_excel,
            pi_sheet=args.pi_sheet,
            aoi_rect=aoi_rect_for_task(task, args),
        )

        n_files += 1

    print(f"[DONE] Batch processed {n_files} eye-tracking.csv files under {base_path}")


### Code for SA (Simulated annealing)
def process_pi_for_sa(
    window_size: float,
    context: str,
    *,
    base_path: str,
    stride: Optional[float] = None,
    screen_width_px: float = DEFAULT_SCREEN_WIDTH,
    screen_height_px: float = DEFAULT_SCREEN_HEIGHT,
    timezone_name: str = "Asia/Seoul",
    pi_excel: Optional[str] = None,
    pi_sheet: str = "eye-tracking",
    task_time_json: Optional[str] = None,
    only_client_id: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    In-memory batch PI extraction (NO FILE SAVE).

    - Forces window_size (and stride) for vst/ast/gng/flanker/dnb to the given values
    - Processes ONLY the specified context (task), e.g. "ast"
    - Returns a list of output JSON objects (same structure as file output), one per processed CSV.

    Returns:
      List of out_obj dicts, each like:
        {
          "client_id": ...,
          "context": ...,
          "sensor_modality": "eye-tracking",
          "screen": {...},
          "config": {"window_size": ..., "stride": ...},
          "windows": [...]
        }
    """
    ctx = str(context).lower().strip()
    allowed = {
        "vst", "ast", "gng", "flanker", "dnb",
        "rocf_copy_low", "rocf_recall_low",
        "rocf_copy_mid", "rocf_recall_mid",
        "rocf_copy_high", "rocf_recall_high",
    }
    if ctx not in allowed:
        raise ValueError(f"Unknown context/task: {context!r}. Expected one of: {sorted(allowed)}")

    stride_val = float(window_size if stride is None else stride)

    # --- Backup globals so we can restore after running ---
    global TASK_WINDOW_SIZE_SEC, TASK_STRIDE_SEC
    _old_ws = dict(TASK_WINDOW_SIZE_SEC)
    _old_st = dict(TASK_STRIDE_SEC)

    def _process_one_file_to_obj(
        input_csv: str,
        client_id: str,
        task: str,
        task_start_ms: Optional[int],
        task_end_ms: Optional[int],
        window_size_s: float,
        stride_s: float,
    ) -> Dict[str, Any]:
        """
        A no-save version of process_one_file(): builds and RETURNS out_obj.
        """
        try:
            df = pd.read_csv(input_csv)
        except EmptyDataError:
            # mimic original behavior: skip
            return {}

        if df.empty:
            return {}

        for col in ("timestamp", "x", "y"):
            if col not in df.columns:
                raise ValueError(f"[{input_csv}] Missing required column: {col}")

        # Normalize timestamp column to ms int
        df["timestamp"] = parse_timestamp_column(df["timestamp"])
        df = df.sort_values("timestamp").reset_index(drop=True)

        pi_fullnames = load_pi_fullnames_from_excel(pi_excel, sheet_name=pi_sheet) if pi_excel else None

        window_ms = int(round(window_size_s * 1000.0))
        stride_ms = int(round(stride_s * 1000.0))

        # Window anchor: task interval if provided else data-driven
        if task_start_ms is not None and task_end_ms is not None and int(task_end_ms) > int(task_start_ms):
            t_min = int(task_start_ms)
            t_max = int(task_end_ms)
        else:
            t_min = int(df["timestamp"].min())
            t_max = int(df["timestamp"].max())

        windows = iter_windows_partial(t_min, t_max, window_ms, stride_ms)

        out_obj: Dict[str, Any] = {
            "client_id": client_id,
            "context": task,
            "sensor_modality": "eye-tracking",
            "screen": {"width": float(screen_width_px), "height": float(screen_height_px)},
            "config": {"window_size": float(window_size_s), "stride": float(stride_s)},
            "windows": [],
        }

        for wnum, w_start, w_end in windows:
            dfw = df[(df["timestamp"] >= w_start) & (df["timestamp"] < w_end)]

            base_vecs = compute_window_base_vectors(
                dfw=dfw,
                w_start_ms=w_start,
                w_end_ms=w_end,
                screen_w=float(screen_width_px),
                screen_h=float(screen_height_px),
            )
            pi_dict = (
                build_pi_from_fullnames(base_vecs, pi_fullnames)
                if pi_fullnames is not None
                else build_pi_from_base_vectors(base_vecs)
            )
            pi_dict.update(build_direct_aoi_pi(base_vecs))

            out_obj["windows"].append(
                {
                    "window_number": int(wnum),
                    "start_timestamp": ms_to_iso(int(w_start), tz_name=timezone_name),
                    "end_timestamp": ms_to_iso(int(w_end), tz_name=timezone_name),
                    "pi": pi_dict,
                }
            )

        return out_obj

    try:
        # --- Override window/stride ONLY for the 5 main tasks ---
        for t in ("vst", "ast", "gng", "flanker", "dnb"):
            TASK_WINDOW_SIZE_SEC[t] = float(window_size)
            TASK_STRIDE_SEC[t] = float(stride_val)

        base_path_abs = os.path.abspath(base_path)
        client_filter = set(only_client_id) if only_client_id else None

        # Optional: load task interval index
        task_time_index = None
        if task_time_json:
            try:
                task_time_index = load_task_time_index(task_time_json)
            except Exception:
                task_time_index = None

        results: List[Dict[str, Any]] = []

        for root, dirs, files in os.walk(base_path_abs):
            if "eye-tracking.csv" not in files:
                continue

            csv_path = os.path.join(root, "eye-tracking.csv")
            rel = os.path.relpath(csv_path, base_path_abs)
            parts = rel.split(os.sep)

            # expected: date/client_id/task/eye-tracking/archive_id/eye-tracking.csv
            if len(parts) < 6:
                continue

            client_id = parts[1]
            task = parts[2]
            modality_dir = parts[3]

            if modality_dir != "eye-tracking":
                continue

            # Only requested context
            if str(task).lower() != ctx:
                continue

            # optional client filter
            if client_filter is not None and client_id not in client_filter:
                continue

            window_size_sec = resolve_window_size_for_task(task, float(window_size))
            stride_sec = resolve_stride_for_task(task, float(stride_val))

            task_start_ms = None
            task_end_ms = None
            if task_time_index is not None:
                key = (str(client_id), str(task))
                if key in task_time_index:
                    task_start_ms, task_end_ms = task_time_index[key]

            out_obj = _process_one_file_to_obj(
                input_csv=csv_path,
                client_id=client_id,
                task=task,
                task_start_ms=task_start_ms,
                task_end_ms=task_end_ms,
                window_size_s=float(window_size_sec),
                stride_s=float(stride_sec),
            )

            # skip empties
            if out_obj and out_obj.get("windows") is not None:
                results.append(out_obj)

        return results

    finally:
        # Restore original dicts so importing this module doesn't permanently mutate behavior
        TASK_WINDOW_SIZE_SEC = _old_ws
        TASK_STRIDE_SEC = _old_st



def main() -> None:
    ap = argparse.ArgumentParser()

    # (A) 단일 파일 모드용
    ap.add_argument("--input_csv", help="Single input CSV (eye-tracking)")
    ap.add_argument("--output_json", help="Single output JSON path")

    # (B) 배치 모드용
    ap.add_argument(
        "--base_path",
        default=str(DEFAULT_RAW_DATA_DIR),
        help="Base path containing date/client/task/.../eye-tracking.csv. Default: <project_root>/pdss_data",
    )
    ap.add_argument(
        "--output_base_path",
        default=str(DEFAULT_PI_OUTPUT_DIR),
        help="Base path to write JSONs into. Default: <project_root>/data_preprocessing/primitive_indicator",
    )

    ap.add_argument("--window_size_sec", type=float, default=DEFAULT_WINDOW_SIZE_S, help="seconds")
    ap.add_argument("--stride_sec", type=float, default=DEFAULT_STRIDE_S, help="seconds")

    # 단일 모드에서만 직접 주입하는 메타데이터 (배치 모드에서는 경로에서 뽑음)
    ap.add_argument("--client_id", default="client_001")
    ap.add_argument("--context", default="TASK_X")
    ap.add_argument("--sensor_modality", default="eye-tracking")
    ap.add_argument("--timezone", default="Asia/Seoul")

    ap.add_argument("--screen_width_px", type=float, default=DEFAULT_SCREEN_WIDTH)
    ap.add_argument("--screen_height_px", type=float, default=DEFAULT_SCREEN_HEIGHT)
    ap.add_argument("--aoi_x_min", type=float, default=DEFAULT_AOI_X_MIN)
    ap.add_argument("--aoi_x_max", type=float, default=DEFAULT_AOI_X_MAX)
    ap.add_argument("--aoi_y_min", type=float, default=DEFAULT_AOI_Y_MIN)
    ap.add_argument("--aoi_y_max", type=float, default=DEFAULT_AOI_Y_MAX)
    ap.add_argument(
        "--aoi_tasks",
        nargs="*",
        default=list(DEFAULT_AOI_TASKS),
        help="Tasks that should use AOI extraction. Default: gng flanker dnb",
    )
    
    ap.add_argument("--pi_excel", help="Excel file path containing PI 'Name' column")
    ap.add_argument("--pi_sheet", default="eye-tracking", help="Sheet name in the PI excel")

    ap.add_argument("--task_time_json", help="JSON containing start_t/end_t per (client_id,task) to anchor windowing")


    # ★ 특정 client만 돌리고 싶을 때: --only_client_id clientA clientB ...
    ap.add_argument(
        "--only_client_id",
        nargs="*",
        help="If given, only process these client_ids in batch mode",
    )

    args = ap.parse_args()

    # --- 모드 결정 ---

    # 1) 단일 파일 모드: input_csv + output_json 이 둘 다 주어지면 이 모드
    if args.input_csv and args.output_json:
        task_start_ms = None
        task_end_ms = None
        if getattr(args, "task_time_json", None):
            try:
                _idx = load_task_time_index(args.task_time_json)
                key = (str(args.client_id), str(args.context))
                if key in _idx:
                    task_start_ms, task_end_ms = _idx[key]
            except Exception as e:
                print(f"[WARN] Failed to load task_time_json in single-file mode: {args.task_time_json} ({e}). Falling back to data-driven windowing.")

        process_one_file(
            input_csv=args.input_csv,
            output_json=args.output_json,
            client_id=args.client_id,
            context=args.context,
            sensor_modality=args.sensor_modality,
            timezone_name=args.timezone,
            screen_w=float(args.screen_width_px),
            screen_h=float(args.screen_height_px),
            window_size_s=args.window_size_sec,
            stride_s=args.stride_sec,
            task_start_ms=task_start_ms,
            task_end_ms=task_end_ms,
            pi_excel=args.pi_excel,
            pi_sheet=args.pi_sheet,
            aoi_rect=aoi_rect_for_task(args.context, args),
        )
        return

    # 2) 배치 모드: base_path + output_base_path 가 있어야 함
    if args.base_path and args.output_base_path:
        run_batch_mode(args)
        return

    # 둘 다 아니면 에러
    raise SystemExit(
        "You must either:\n"
        "  (a) provide --input_csv and --output_json for single-file mode, OR\n"
        "  (b) provide --base_path and --output_base_path for batch mode."
    )


if __name__ == "__main__":
    main()
